"""
Excel 匯出服務
提供資料匯出至 Excel 的功能，包含格式化與樣式設定
"""

import os
import sys
import logging
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy
import zipfile
import tempfile
import shutil


class ExcelExporter:
    """Excel 匯出服務"""
    
    def __init__(self):
        """初始化 Excel 匯出器"""
        self.logger = logging.getLogger("ICRLogger")
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def export_to_excel(
        self,
        scored_rows: List[Dict[str, any]],
        output_columns: List[str],
        output_path: str,
        answer_data: List[Dict] = None,
        base_columns: List[str] = None,
        doc_type: str = None
    ) -> bool:
        """
        將評分後的資料匯出至 Excel
        
        Args:
            scored_rows: 評分後的資料列表
            output_columns: 輸出欄位列表
            output_path: 輸出檔案路徑
            answer_data: 答案資料（用於計算統計）
            base_columns: 基礎欄位列表（用於計算統計）
            doc_type: 文件類型（ARC/Health/Employment）
            
        Returns:
            成功返回 True，失敗返回 False
        """
        self.logger.info(f"開始匯出 Excel - 路徑: {output_path}")
        
        if not scored_rows:
            self.logger.warning("沒有資料可以匯出")
            return False
        
        try:
            # 如果有答案資料且需要樞紐分析，從模板載入；否則建立新工作簿
            template_path = None
            if answer_data and base_columns:
                template_path = self._find_template_file(doc_type)
            
            if template_path:
                self.logger.info(f"從模板載入工作簿: {template_path}")
                wb = load_workbook(template_path)
                
                # 刪除 Result 工作表（如果存在）並重新創建
                if 'Result' in wb.sheetnames:
                    del wb['Result']
                    self.logger.info("已刪除舊的 Result 工作表")
                
                # 創建新的 Result 工作表並移到第一個位置
                ws = wb.create_sheet('Result', 0)
                self.logger.info("已創建新的 Result 工作表")
                
                # 刪除其他可能存在的舊數據工作表
                for sheet_name in ['Statistics', 'Report', 'Analyze']:
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]
            else:
                # 建立新工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = 'Result'
            
            # 寫入標題列
            ws.append(output_columns)
            for cell in ws[1]:
                cell.alignment = self.center_alignment
            
            # 寫入資料列
            for row_data in scored_rows:
                row_values = [row_data.get(col, '') for col in output_columns]
                ws.append(row_values)
            
            # 把 Result 工作表的資料範圍包裝成名為 ResultTable 的 Excel 表格
            try:
                max_row = ws.max_row
                max_col = len(output_columns)
                last_col_letter = get_column_letter(max_col)
                table_ref = f"A1:{last_col_letter}{max_row}"
                
                # 創建新表格（表格名稱必須為 ResultTable 以配合樞紐分析表）
                table = Table(displayName="ResultTable", ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                ws.add_table(table)
                self.logger.info(f"成功創建表格 ResultTable，範圍: {table_ref}")
            except Exception as e:
                self.logger.warning(f"建立 ResultTable 時發生錯誤: {e}")

            # 套用格式
            self._apply_formatting(ws, output_columns)
            self._auto_adjust_columns(ws, output_columns)
            
            # 如果有答案資料，創建統計 Sheet
            if answer_data and base_columns:
                statistics_data = self._create_statistics_sheet(wb, scored_rows, answer_data, base_columns)
                self._create_report_sheet(wb, scored_rows, statistics_data)
                self._create_analyze_sheet(wb, scored_rows, answer_data, base_columns)
                
                # 如果從模板載入，設定樞紐分析快取自動更新
                if template_path:
                    self._set_pivot_refresh_on_load(wb)
                    # 嘗試將 PivotChart 移到第 5 個工作表（index 4）
                    try:
                        if 'PivotChart' in wb.sheetnames:
                            pivot_ws = wb['PivotChart']
                            # 先移除再插入到指定位置
                            wb._sheets.remove(pivot_ws)
                            insert_index = min(4, len(wb._sheets))
                            wb._sheets.insert(insert_index, pivot_ws)
                            self.logger.info('已將 PivotChart 移到第5個工作表')
                        else:
                            self.logger.warning('模板中找不到 PivotChart 工作表，無法移動')
                    except Exception as e:
                        self.logger.warning(f'移動 PivotChart 時發生錯誤: {e}')
            
            # 建立輸出目錄並儲存
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # 如果從模板載入，需要先儲存再修改 XML
            if template_path:
                temp_path = output_path + ".tmp"
                wb.save(temp_path)
                # 修改 XML 設定樞紐分析自動更新
                self._set_pivot_refresh_in_xml(temp_path, output_path)
                self.logger.info(f"Excel 匯出成功（含樞紐分析）: {output_path}，共 {len(scored_rows)} 筆資料")
            else:
                wb.save(output_path)
                self.logger.info(f"Excel 匯出成功: {output_path}，共 {len(scored_rows)} 筆資料")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Excel 匯出失敗: {e}")
            return False
    
    def _apply_formatting(self, ws, output_columns: List[str]) -> None:
        """
        套用儲存格格式
        
        Args:
            ws: Worksheet 物件
            output_columns: 輸出欄位列表
        """
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = self.center_alignment
    
    def _auto_adjust_columns(self, ws, output_columns: List[str]) -> None:
        """
        自動調整欄寬
        
        Args:
            ws: Worksheet 物件
            output_columns: 輸出欄位列表
        """
        for idx, col_name in enumerate(output_columns, start=1):
            column_letter = ws.cell(row=1, column=idx).column_letter
            max_length = len(str(col_name))
            
            # 計算最大長度
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    try:
                        cell_value = str(cell.value) if cell.value else ''
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
            
            # 調整欄寬 (限制在 10-50 之間)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_statistics_sheet(
        self,
        wb,
        scored_rows: List[Dict],
        answer_data: List[Dict],
        base_columns: List[str]
    ) -> List[Dict]:
        """
        創建統計資料 Sheet
        
        Args:
            wb: Workbook 物件
            scored_rows: 評分後的資料
            answer_data: 答案資料
            base_columns: 基礎欄位列表
            
        Returns:
            統計數據列表
        """
        ws = wb.create_sheet(title='Statistics')
        
        # 統計欄位
        stat_columns = [
            "檔名",
            "正確欄位數",
            "實際應該有的項目數",
            "模型輸出的項目數",
            "拿來比較的項目數",
            "Precision",
            "Recall", 
            "F1 Score",
            "Item Accuracy"
        ]
        
        # 寫入標題
        ws.append(stat_columns)
        for cell in ws[1]:
            cell.alignment = self.center_alignment
            
        # 儲存統計數據用於 Report Sheet
        statistics_data = []
        
        # 建立答案字典
        answer_dict = {row.get('檔名', ''): row for row in answer_data if '檔名' in row}
        
        # 定義要排除的欄位
        excluded_fields = {'資料序號', '檔名', '文件類型', '資料類型'}
        
        # 為每個檔案計算統計
        for row in scored_rows:
            file_name = row.get('檔名', '')
            if not file_name or file_name not in answer_dict:
                continue
            
            answer_row = answer_dict[file_name]
            
            # 1. 正確欄位數：只有「答案有值」且「模型有值」且 PASS 才計算
            correct_count = 0
            for field in base_columns:
                if field in excluded_fields:
                    continue
                answer_value = str(answer_row.get(field, '')).strip()
                model_value = str(row.get(field, '')).strip()
                answer_field = f"{field}_答案"
                # 移除括號內的內容（答案）
                if '(' in model_value and ')' in model_value:
                    model_value = model_value.split('(')[0].strip()
                if answer_value and model_value and row.get(answer_field) == 'PASS':
                    correct_count += 1

            # 2. 實際應該有的項目數：答案中非空欄位數（不包含資料序號、檔名、文件類型）
            expected_count = 0
            for field in base_columns:
                if field in excluded_fields:
                    continue
                answer_value = str(answer_row.get(field, '')).strip()
                if answer_value:
                    expected_count += 1

            # 3. 模型輸出的項目數：模型有輸出值的欄位數（不為空且不是 N/A）
            model_output_count = 0
            compared_count = 0  # 拿來比較的項目數：答案有值且模型也有值的欄位數
            
            for field in base_columns:
                if field in excluded_fields:
                    continue
                model_value = str(row.get(field, '')).strip()
                answer_value = str(answer_row.get(field, '')).strip()
                if '(' in model_value and ')' in model_value:
                    model_value = model_value.split('(')[0].strip()
                
                # 模型輸出數：只要模型有值且不是 N/A 就計入
                if model_value and model_value.upper() != 'N/A':
                    model_output_count += 1
                
                # 比較項目數：答案有值且模型也有值才計入
                if answer_value and model_value and model_value.upper() != 'N/A':
                    compared_count += 1
                
                # 調試：打印字段信息
                if model_value or answer_value:
                    self.logger.debug(f"  [{field}] 答案:{answer_value!r} | 模型:{model_value!r} | 模型有值:{bool(model_value and model_value.upper() != 'N/A')} | 可比較:{bool(answer_value and model_value and model_value.upper() != 'N/A')}")
            
            # 計算額外的統計指標
            precision = correct_count / model_output_count if model_output_count > 0 else 0
            recall = correct_count / expected_count if expected_count > 0 else 0
            f1_score = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
            item_accuracy = correct_count / compared_count if compared_count > 0 else 0
            
            # 寫入統計資料
            stat_row = [
                file_name,
                correct_count,
                expected_count,
                model_output_count,
                compared_count,
                f"{precision:.2f}",
                f"{recall:.2f}",
                f"{f1_score:.2f}",
                f"{item_accuracy:.2f}"
            ]
            ws.append(stat_row)
            
            # 儲存統計數據
            statistics_data.append({
                'file_name': file_name,
                'correct_count': correct_count,
                'expected_count': expected_count,
                'model_output_count': model_output_count,
                'compared_count': compared_count,
                'precision': precision,
                'recall': recall,
                'f1_score': f1_score,
                'item_accuracy': item_accuracy
            })
        
        # 套用格式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = self.center_alignment
        
        # 自動調整欄寬
        for idx, col_name in enumerate(stat_columns, start=1):
            column_letter = ws.cell(row=1, column=idx).column_letter
            max_length = len(str(col_name))
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    try:
                        cell_value = str(cell.value) if cell.value else ''
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
            
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return statistics_data
    
    def _create_report_sheet(self, wb: Workbook, scored_rows: List[Dict], statistics_data: List[Dict]) -> None:
        """
        創建報告工作表
        
        Args:
            wb: Workbook 物件
            scored_rows: 評分後的資料
            statistics_data: 統計數據列表
        """
        ws = wb.create_sheet(title='Report')
        
        # 標題行
        ws.append(["分類", "指標", "數值"])
        for cell in ws[1]:
            cell.alignment = self.center_alignment
        
        # 計算統計數據
        total_records = len(scored_rows)
        success_count = total_records  # 成功處理數量
        fail_count = 0  # 處理失敗（先掛0）
        
        # 計算完全正確數量（辨識結果為 PASS）
        pass_count = sum(1 for row in scored_rows if row.get('辨識結果') == 'PASS')
        
        # 完全正確率
        perfect_rate = (pass_count / total_records * 100) if total_records > 0 else 0
        
        # 計算平均指標（從 Statistics sheet 的數據計算）
        total_precision = 0
        total_recall = 0
        total_f1 = 0
        total_item_acc = 0
        
        for stat in statistics_data:
            total_precision += stat['precision']
            total_recall += stat['recall']
            total_f1 += stat['f1_score']
            total_item_acc += stat['item_accuracy']
        
        avg_precision = (total_precision / success_count) if success_count > 0 else 0
        avg_recall = (total_recall / success_count) if success_count > 0 else 0
        avg_f1 = (total_f1 / success_count) if success_count > 0 else 0
        avg_item_acc = (total_item_acc / success_count) if success_count > 0 else 0
        avg_char_acc = 0  # 平均字元正確率（先給0）
        
        # 寫入數據行
        report_data = [
            ["基本統計", "總記錄數", total_records],
            ["基本統計", "成功處理", success_count],
            ["基本統計", "處理失敗", fail_count],
            ["基本統計", "完全正確", pass_count],
            ["基本統計", "完全正確率", f"{perfect_rate:.3f}%"],
            ["整體指標", "平均 Precision", f"{avg_precision:.3f}"],
            ["整體指標", "平均 Recall", f"{avg_recall:.3f}"],
            ["整體指標", "平均 F1-Score", f"{avg_f1:.3f}"],
            ["整體指標", "平均項目正確率", f"{avg_item_acc:.3f}"],
            ["整體指標", "平均字元正確率", f"{avg_char_acc:.3f}"]
        ]
        
        for row_data in report_data:
            ws.append(row_data)
        
        # 套用格式
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = self.center_alignment
        
        # 自動調整欄寬
        for idx in range(1, 4):
            column_letter = ws.cell(row=1, column=idx).column_letter
            max_length = 10
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    try:
                        cell_value = str(cell.value) if cell.value else ''
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
            
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_analyze_sheet(
        self,
        wb: Workbook,
        scored_rows: List[Dict],
        answer_data: List[Dict],
        base_columns: List[str]
    ) -> None:
        """
        創建欄位分析工作表
        
        Args:
            wb: Workbook 物件
            scored_rows: 評分後的資料
            answer_data: 答案資料
            base_columns: 基礎欄位列表
        """
        ws = wb.create_sheet(title='Analyze')
        
        # 標題行
        analyze_columns = [
            "欄位名稱",
            "總出現次數",
            "完全正確",
            "部分正確",
            "完全錯誤",
            "缺失",
            "多餘",
            "正確率",
            "錯誤率",
            "部分正確率",
            "模式"
        ]
        ws.append(analyze_columns)
        for cell in ws[1]:
            cell.alignment = self.center_alignment
        
        # 排除的欄位
        excluded_fields = {'檔名', '資料序號', '資料類型', 'file_name', 'data_id', 'doc_type'}
        
        # 分析每個欄位
        for field in base_columns:
            if field in excluded_fields:
                continue
            
            # 1. 總出現次數：答案表裡該欄位有值的數量
            total_count = 0
            for answer_row in answer_data:
                answer_value = str(answer_row.get(field, '')).strip()
                if answer_value:
                    total_count += 1
            
            # 如果該欄位在答案中從未出現，跳過
            if total_count == 0:
                continue
            
            # 2. 完全正確：答案有值且該欄位_答案為 PASS 的數量
            correct_count = 0
            for row in scored_rows:
                # 找到對應的答案行
                file_name = row.get('檔名', '')
                answer_row = None
                for ans in answer_data:
                    if str(ans.get('檔名', '')) == file_name:
                        answer_row = ans
                        break
                
                if answer_row:
                    answer_value = str(answer_row.get(field, '')).strip()
                    result_field = f"{field}_答案"
                    # 答案有值且評分為 PASS
                    if answer_value and row.get(result_field) == 'PASS':
                        correct_count += 1
            
            # 3. 部分正確（先給0）
            partial_correct = 0
            
            # 4. 完全錯誤：該欄位_答案為 FAIL 的數量
            fail_count = 0
            for row in scored_rows:
                result_field = f"{field}_答案"
                if row.get(result_field) == 'FAIL':
                    fail_count += 1
            
            # 5. 缺失：答案有值但模型輸出為空值(N/A)的數量
            missing_count = 0
            for row in scored_rows:
                # 找到對應的答案行
                file_name = row.get('檔名', '')
                answer_row = None
                for ans in answer_data:
                    if str(ans.get('檔名', '')) == file_name:
                        answer_row = ans
                        break
                
                if answer_row:
                    answer_value = str(answer_row.get(field, '')).strip()
                    model_value = str(row.get(field, '')).strip()
                    # 移除括號內容
                    if '(' in model_value and ')' in model_value:
                        model_value = model_value.split('(')[0].strip()
                    
                    # 答案有值但模型輸出為空或 N/A
                    if answer_value and (not model_value or model_value.upper() == 'N/A'):
                        missing_count += 1
            
            # 6. 多餘：答案表無值但模型輸出有值的數量
            extra_count = 0
            for i, row in enumerate(scored_rows):
                # 找到對應的答案行
                file_name = row.get('檔名', '')
                answer_row = None
                for ans in answer_data:
                    if str(ans.get('檔名', '')) == file_name:
                        answer_row = ans
                        break
                
                if answer_row:
                    answer_value = str(answer_row.get(field, '')).strip()
                    model_value = str(row.get(field, '')).strip()
                    # 移除括號內容
                    if '(' in model_value and ')' in model_value:
                        model_value = model_value.split('(')[0].strip()
                    
                    # 答案無值但模型有值
                    if not answer_value and model_value and model_value.upper() != 'N/A':
                        extra_count += 1
            
            # 7. 正確率：完全正確/總出現次數
            accuracy = (correct_count / total_count * 100) if total_count > 0 else 0
            
            # 8. 錯誤率：完全錯誤/總出現次數
            error_rate = (fail_count / total_count * 100) if total_count > 0 else 0
            
            # 9. 部分正確率（先給0）
            partial_rate = 0
            
            # 10. 模式
            mode = "嚴謹"
            
            # 寫入數據行
            analyze_row = [
                field,
                total_count,
                correct_count,
                partial_correct,
                fail_count,
                missing_count,
                extra_count,
                f"{accuracy:.3f}%",
                f"{error_rate:.3f}%",
                f"{partial_rate:.3f}%",
                mode
            ]
            ws.append(analyze_row)
        
        # 套用格式
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = self.center_alignment
        
        # 自動調整欄寬
        for idx in range(1, len(analyze_columns) + 1):
            column_letter = ws.cell(row=1, column=idx).column_letter
            max_length = 10
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    try:
                        cell_value = str(cell.value) if cell.value else ''
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
            
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    

    def _find_template_file(self, doc_type: str = None) -> str:
        """
        在專案根目錄的 excel_template 資料夾中搜尋模板檔案
        
        Args:
            doc_type: 文件類型（ARC/Health/Employment）
            
        Returns:
            模板檔案路徑，找不到則返回 None
        """
        # 1. 定義模板對照表
        template_mapping = {
            'ARC': 'ARC_Sample.xlsx',
            'Health': 'Health_Sample.xlsx',
            'Employment': 'Employment_Sample.xlsx'
        }
        
        if doc_type not in template_mapping:
            self.logger.warning(f"未知的文件類型: {doc_type}，無法選擇模板")
            return None

        # 2. 獲取專案根目錄或 exe 所在目錄
        # 支援開發模式和打包後的執行模式
        if getattr(sys, 'frozen', False):
            # 打包後：exe 所在目錄
            project_root = os.path.dirname(sys.executable)
        else:
            # 開發模式：專案根目錄
            current_file_path = os.path.abspath(__file__)
            project_root = os.path.dirname(os.path.dirname(current_file_path))
        
        # 3. 組合出目標路徑：根目錄 / excel_template / 檔名
        template_dir = os.path.join(project_root, 'excel_template')
        template_path = os.path.join(template_dir, template_mapping[doc_type])

        # 4. 檢查檔案是否存在
        if os.path.exists(template_path):
            self.logger.info(f"找到模板檔案: {template_path}")
            return template_path
        
        self.logger.warning(f"找不到模板檔案: {template_path}")
        return None
    
    def _set_pivot_refresh_on_load(self, wb: Workbook) -> None:
        """
        設定樞紐分析表在開啟檔案時自動更新
        (此方法已廢棄，改用 _set_pivot_refresh_in_xml)
        
        Args:
            wb: Workbook 物件
        """
        pass
    
    def _set_pivot_refresh_in_xml(self, temp_path: str, final_path: str) -> None:
        """
        透過修改 Excel XML 來設定樞紐分析快取自動更新
        
        Args:
            temp_path: 臨時檔案路徑
            final_path: 最終輸出路徑
        """
        try:
            # Excel 檔案本質上是 ZIP 壓縮檔
            temp_dir = tempfile.mkdtemp()
            
            # 解壓縮 Excel 檔案
            with zipfile.ZipFile(temp_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # 修改所有樞紐分析快取 XML
            pivot_cache_dir = os.path.join(temp_dir, 'xl', 'pivotCache')
            if os.path.exists(pivot_cache_dir):
                modified_count = 0
                for filename in os.listdir(pivot_cache_dir):
                    if filename.startswith('pivotCacheDefinition') and filename.endswith('.xml'):
                        file_path = os.path.join(pivot_cache_dir, filename)
                        
                        # 讀取 XML
                        with open(file_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                        
                        # 修改 refreshOnLoad 屬性
                        # 處理已有 refreshOnLoad 的情況
                        if 'refreshOnLoad=' in content:
                            content = content.replace('refreshOnLoad="0"', 'refreshOnLoad="1"')
                            content = content.replace('refreshOnLoad="false"', 'refreshOnLoad="1"')
                        else:
                            # 在 pivotCacheDefinition 標籤中加入 refreshOnLoad="1"
                            content = content.replace(
                                '<pivotCacheDefinition',
                                '<pivotCacheDefinition refreshOnLoad="1"',
                                1
                            )
                        
                        # 寫回檔案
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(content)
                        
                        modified_count += 1
                        self.logger.info(f"已修改樞紐分析快取定義: {filename}")
                
                if modified_count > 0:
                    self.logger.info(f"共修改 {modified_count} 個樞紐分析快取設定")
            
            # 重新壓縮成 Excel 檔案
            with zipfile.ZipFile(final_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        zip_ref.write(file_path, arcname)
            
            # 清理臨時檔案
            shutil.rmtree(temp_dir)
            os.remove(temp_path)
            
            self.logger.info("樞紐分析快取已設定為開啟時自動更新")
            
        except Exception as e:
            self.logger.error(f"設定樞紐分析自動更新時發生錯誤: {e}")
            # 如果失敗，至少保留臨時檔案作為輸出
            if os.path.exists(temp_path):
                shutil.move(temp_path, final_path)

