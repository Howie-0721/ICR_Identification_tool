"""
檔案處理輔助函式
提供 CSV 和 Excel 檔案的讀取功能
"""

import csv
import logging
from typing import List, Dict
from openpyxl import load_workbook


def read_csv_data(file_path: str) -> List[Dict[str, str]]:
    """
    讀取 CSV 文件
    
    Args:
        file_path: CSV 檔案路徑
        
    Returns:
        資料列表 (每列為一個字典)
        
    Raises:
        Exception: 檔案讀取失敗時拋出
    """
    logger = logging.getLogger("ICRLogger")
    logger.debug(f"讀取 CSV 文件: {file_path}")
    
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            data = list(csv.DictReader(f))
        logger.debug(f"成功讀取 {len(data)} 筆 CSV 資料")
        return data
    except Exception as e:
        logger.error(f"讀取 CSV 失敗: {e}")
        raise


def read_excel_data(file_path: str) -> List[Dict[str, str]]:
    """
    讀取 Excel 文件，過濾空行並處理數值格式
    
    Args:
        file_path: Excel 檔案路徑
        
    Returns:
        資料列表 (每列為一個字典)
        
    Raises:
        Exception: 檔案讀取失敗時拋出
    """
    logger = logging.getLogger("ICRLogger")
    logger.debug(f"讀取 Excel 文件: {file_path}")
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 讀取標題列
        headers = [cell.value for cell in ws[1]]
        
        # 讀取資料列
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            all_empty = True
            
            for header, value in zip(headers, row):
                # 處理數值格式 (將浮點數轉為整數字串)
                if isinstance(value, float) and value == int(value):
                    val = str(int(value))
                else:
                    val = str(value or '').strip()
                
                row_dict[header] = val
                
                if val:
                    all_empty = False
            
            # 過濾空行
            if not all_empty:
                data.append(row_dict)
        
        logger.debug(f"成功讀取 {len(data)} 筆 Excel 資料")
        return data
        
    except Exception as e:
        logger.error(f"讀取 Excel 失敗: {e}")
        raise
